import openpyxl
import json

import os
# Excel 和脚本在同一目录
script_dir = os.path.dirname(os.path.abspath(__file__))
xlsx_path = os.path.join(script_dir, "标注二级分类后的方剂表.xlsx")
wb = openpyxl.load_workbook(xlsx_path, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n===== 工作表: {sheet_name} =====")
    print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
    
    # 打印前5行看看表头和数据
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=False), 1):
        vals = []
        for cell in row:
            vals.append(f"{cell.column_letter}:{cell.value}")
        print(f"  第{row_idx}行: {vals}")
    
    # 打印所有列名（第一行）
    print("\n  列名:")
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        print(f"    列{col}: {header}")
