"""
从 标注二级分类后的方剂表.xlsx 生成 SQL 数据导入脚本
输出: herb_data.sql, formula_data.sql, symptom_dict_data.sql
"""
import openpyxl
import hashlib
import sys
import io

# 解决 GBK 编码问题
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_PATH = r"c:\Users\maoto\Desktop\专业相关\大创\中医医疗页面\页面代码\project-main\1\medicine (1) (1)\medicine\medicine\标注二级分类后的方剂表.xlsx"
OUTPUT_DIR = r"c:\Users\maoto\Desktop\专业相关\大创\中医医疗页面\页面代码\project-main\1\medicine (1) (1)\medicine\medicine\database"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['Sheet1']

# ============ 1. 提取所有唯一方剂和药材 ============
formulas = {}       # key: formula_name -> {name, category_l1, category_l2, functions, indications, modifications, source}
herbs_in_excel = set()  # set of herb names
formula_herbs = []  # list of {formula_name, herb_name, dosage, sort_order}

for row_idx in range(4, ws.max_row + 1):  # 从第4行开始（前3行是说明）
    vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 16)]
    formula_name = vals[8]  # I列: 方剂名
    herb_name = vals[9]     # J列: 药材名
    content = vals[3]       # D列: 含量 (e.g., "12g")
    func = vals[4]          # E列: 功能
    indications = vals[5]   # F列: 主治
    modifications = vals[6] # G列: 修改
    source = vals[10]       # K列: 来源
    cat_l1 = vals[11]       # L列: 一级方剂类型
    cat_l2 = vals[12]       # M列: 二级方剂类型
    origin = vals[14]       # O列: 出处
    
    if not formula_name or not herb_name:
        continue
    
    herb_name = str(herb_name).strip()
    formula_name = str(formula_name).strip()
    
    # 过滤乱码字符（如 U+FFFC OBJ 替换字符）
    if '\ufffc' in herb_name or len(herb_name) < 2:
        continue
    
    herbs_in_excel.add(herb_name)
    
    # 收集方剂信息
    if formula_name not in formulas:
        formulas[formula_name] = {
            'name': formula_name,
            'category_l1': str(cat_l1).strip() if cat_l1 else '',
            'category_l2': str(cat_l2).strip() if cat_l2 else '',
            'functions': str(func).strip() if func else '',
            'indications': str(indications).strip() if indications else '',
            'modifications': str(modifications).strip() if modifications else '',
            'source': str(source or origin or '').strip()
        }
    
    # 收集方剂 - 药材关联（去重 + 防乱码）
    clean_hb = str(herb_name).strip() if herb_name else ''
    if not clean_hb or len(clean_hb) < 2 or '\ufffc' in clean_hb:
        continue
    
    dosage_str = str(content).strip() if content is not None else ''
    formula_herbs.append({
        'formula_name': formula_name,
        'herb_name': clean_hb,
        'dosage': dosage_str
    })
print(f"找到 {len(formulas)} 个方剂")
print(f"找到 {len(herbs_in_excel)} 种药材")
print(f"去重前 {len(formula_herbs)} 条方剂-药材关联")

# 去重 formula_herbs
seen_pairs = set()
unique_fh = []
for fh in formula_herbs:
    key = (fh['formula_name'], fh['herb_name'])
    if key not in seen_pairs:
        seen_pairs.add(key)
        unique_fh.append(fh)
formula_herbs = unique_fh
print(f"去重后 {len(formula_herbs)} 条方剂-药材关联")

# ============ 2. 生成 herb 数据 ============
# 带完整属性的常用中药材数据（从 Excel 中提取 + 补充）
TCM_HERBS = {
    # 解表药
    '麻黄': {'category': '解表药', 'nature': '温', 'taste': '辛、微苦', 'meridian': '肺、膀胱', 'min_dosage': 2, 'max_dosage': 10, 'is_toxic': 0, 'functions': '发汗解表，宣肺平喘，利水消肿', 'indications': '风寒感冒，咳嗽气喘，风水水肿'},
    '桂枝': {'category': '解表药', 'nature': '温', 'taste': '辛、甘', 'meridian': '心、肺、膀胱', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '发汗解肌，温通经脉，助阳化气', 'indications': '风寒感冒，寒凝血滞诸痛，痰饮，蓄水证'},
    '紫苏叶': {'category': '解表药', 'nature': '温', 'taste': '辛', 'meridian': '肺、脾', 'min_dosage': 5, 'max_dosage': 10, 'is_toxic': 0, 'functions': '解表散寒，行气宽中', 'indications': '风寒感冒，脾胃气滞'},
    '生姜': {'category': '解表药', 'nature': '微温', 'taste': '辛', 'meridian': '肺、脾、胃', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '解表散寒，温中止呕，温肺止咳', 'indications': '风寒感冒，脾胃寒证，胃寒呕吐'},
    '防风': {'category': '解表药', 'nature': '微温', 'taste': '辛、甘', 'meridian': '膀胱、肝、脾', 'min_dosage': 5, 'max_dosage': 10, 'is_toxic': 0, 'functions': '祛风解表，胜湿止痛，止痉', 'indications': '外感表证，风疹瘙痒，风湿痹痛'},
    '羌活': {'category': '解表药', 'nature': '温', 'taste': '辛、苦', 'meridian': '膀胱、肾', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '解表散寒，祛风胜湿，止痛', 'indications': '风寒感冒，风寒湿痹'},
    '白芷': {'category': '解表药', 'nature': '温', 'taste': '辛', 'meridian': '肺、胃、大肠', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '解表散寒，祛风止痛，通鼻窍', 'indications': '风寒感冒，头痛，鼻塞'},
    '细辛': {'category': '解表药', 'nature': '温', 'taste': '辛', 'meridian': '肺、肾', 'min_dosage': 1, 'max_dosage': 3, 'is_toxic': 1, 'functions': '解表散寒，祛风止痛，通窍', 'indications': '风寒感冒，头痛，牙痛'},
    '荆芥': {'category': '解表药', 'nature': '微温', 'taste': '辛', 'meridian': '肺、肝', 'min_dosage': 5, 'max_dosage': 10, 'is_toxic': 0, 'functions': '祛风解表，透疹消疮', 'indications': '外感表证，麻疹不透，疮疡初起'},
    '薄荷': {'category': '解表药', 'nature': '凉', 'taste': '辛', 'meridian': '肺、肝', 'min_dosage': 3, 'max_dosage': 6, 'is_toxic': 0, 'functions': '疏散风热，清利头目，利咽透疹', 'indications': '风热感冒，头痛目赤，咽喉肿痛'},
    '牛蒡子': {'category': '解表药', 'nature': '寒', 'taste': '辛、苦', 'meridian': '肺、胃', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '疏散风热，宣肺祛痰，利咽透疹', 'indications': '风热感冒，咳嗽痰多，咽喉肿痛'},
    '蝉蜕': {'category': '解表药', 'nature': '寒', 'taste': '甘', 'meridian': '肺、肝', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '疏散风热，利咽开音，透疹', 'indications': '风热感冒，咽痛音哑，麻疹不透'},
    
    # 清热药
    '石膏': {'category': '清热药', 'nature': '大寒', 'taste': '辛、甘', 'meridian': '肺、胃', 'min_dosage': 15, 'max_dosage': 60, 'is_toxic': 0, 'functions': '清热泻火，除烦止渴', 'indications': '壮热烦渴，肺热喘咳'},
    '知母': {'category': '清热药', 'nature': '寒', 'taste': '苦、甘', 'meridian': '肺、胃、肾', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '清热泻火，滋阴润燥', 'indications': '热病烦渴，肺热燥咳'},
    '黄芩': {'category': '清热药', 'nature': '寒', 'taste': '苦', 'meridian': '肺、胆、脾、胃、大肠、小肠', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '清热燥湿，泻火解毒，止血安胎', 'indications': '湿温暑湿，肺热咳嗽，热毒疮疡'},
    '黄连': {'category': '清热药', 'nature': '寒', 'taste': '苦', 'meridian': '心、脾、胃、胆、大肠', 'min_dosage': 2, 'max_dosage': 10, 'is_toxic': 0, 'functions': '清热燥湿，泻火解毒', 'indications': '湿热痞满，呕吐吞酸，泻痢'},
    '黄柏': {'category': '清热药', 'nature': '寒', 'taste': '苦', 'meridian': '肾、膀胱', 'min_dosage': 3, 'max_dosage': 12, 'is_toxic': 0, 'functions': '清热燥湿，泻火解毒，退虚热', 'indications': '湿热下注，带下黄稠，阴虚发热'},
    '金银花': {'category': '清热药', 'nature': '寒', 'taste': '甘', 'meridian': '肺、心、胃', 'min_dosage': 6, 'max_dosage': 15, 'is_toxic': 0, 'functions': '清热解毒，疏散风热', 'indications': '痈肿疔疮，外感风热，热毒血痢'},
    '连翘': {'category': '清热药', 'nature': '微寒', 'taste': '苦', 'meridian': '肺、心、小肠', 'min_dosage': 6, 'max_dosage': 15, 'is_toxic': 0, 'functions': '清热解毒，消肿散结，疏散风热', 'indications': '痈肿疮毒，瘰疬痰核，风热外感'},
    '生地黄': {'category': '清热药', 'nature': '寒', 'taste': '甘、苦', 'meridian': '心、肝、肾', 'min_dosage': 10, 'max_dosage': 15, 'is_toxic': 0, 'functions': '清热凉血，养阴生津', 'indications': '热入营血，舌绛烦渴，阴虚内热'},
    '玄参': {'category': '清热药', 'nature': '微寒', 'taste': '甘、苦、咸', 'meridian': '肺、胃、肾', 'min_dosage': 10, 'max_dosage': 15, 'is_toxic': 0, 'functions': '清热凉血，泻火解毒，滋阴', 'indications': '温邪入营，烦躁谵语，阴虚火旺'},
    '牡丹皮': {'category': '清热药', 'nature': '微寒', 'taste': '苦、辛', 'meridian': '心、肝、肾', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '清热凉血，活血化瘀', 'indications': '热入营血，温毒发斑，跌打损伤'},
    '赤芍': {'category': '清热药', 'nature': '微寒', 'taste': '苦', 'meridian': '肝', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '清热凉血，散瘀止痛', 'indications': '温毒发斑，目赤肿痛，跌打损伤'},
    '青蒿': {'category': '清热药', 'nature': '寒', 'taste': '苦、辛', 'meridian': '肝、胆', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '清虚热，除骨蒸，解暑截疟', 'indications': '温邪伤阴，阴虚发热，疟疾寒热'},
    
    # 泻下药
    '大黄': {'category': '泻下药', 'nature': '寒', 'taste': '苦', 'meridian': '脾、胃、大肠、肝、心包', 'min_dosage': 3, 'max_dosage': 15, 'is_toxic': 0, 'functions': '泻下攻积，清热泻火，凉血解毒', 'indications': '积滞便秘，热毒疮疡，瘀血证'},
    '芒硝': {'category': '泻下药', 'nature': '寒', 'taste': '咸、苦', 'meridian': '胃、大肠', 'min_dosage': 10, 'max_dosage': 15, 'is_toxic': 0, 'functions': '泻下攻积，润燥软坚，清热消肿', 'indications': '实热积滞，大便燥结'},
    '甘遂': {'category': '泻下药', 'nature': '寒', 'taste': '苦', 'meridian': '肺、肾、大肠', 'min_dosage': 0.5, 'max_dosage': 1, 'is_toxic': 1, 'functions': '泻水逐饮，消肿散结', 'indications': '水肿胀满，胸腹积水'},
    
    # 祛风湿药
    '独活': {'category': '祛风湿药', 'nature': '微温', 'taste': '辛、苦', 'meridian': '肾、膀胱', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '祛风湿，止痛，解表', 'indications': '风寒湿痹，腰膝疼痛'},
    '秦艽': {'category': '祛风湿药', 'nature': '平', 'taste': '辛、苦', 'meridian': '胃、肝、胆', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '祛风湿，通络止痛，退虚热', 'indications': '风湿痹痛，筋脉拘挛'},
    
    # 化湿药
    '厚朴': {'category': '化湿药', 'nature': '温', 'taste': '苦、辛', 'meridian': '脾、胃、肺、大肠', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '燥湿消痰，下气除满', 'indications': '湿阻中焦，食积气滞'},
    '苍术': {'category': '化湿药', 'nature': '温', 'taste': '辛、苦', 'meridian': '脾、胃、肝', 'min_dosage': 5, 'max_dosage': 10, 'is_toxic': 0, 'functions': '燥湿健脾，祛风散寒', 'indications': '湿阻中焦，风湿痹证'},
    '藿香': {'category': '化湿药', 'nature': '微温', 'taste': '辛', 'meridian': '脾、胃、肺', 'min_dosage': 5, 'max_dosage': 10, 'is_toxic': 0, 'functions': '芳香化湿，和中止呕', 'indications': '湿阻中焦，呕吐'},
    
    # 利水渗湿药
    '茯苓': {'category': '利水渗湿药', 'nature': '平', 'taste': '甘、淡', 'meridian': '心、脾、肾', 'min_dosage': 10, 'max_dosage': 15, 'is_toxic': 0, 'functions': '利水渗湿，健脾宁心', 'indications': '水肿尿少，痰饮眩悸，脾虚食少'},
    '泽泻': {'category': '利水渗湿药', 'nature': '寒', 'taste': '甘、淡', 'meridian': '肾、膀胱', 'min_dosage': 6, 'max_dosage': 10, 'is_toxic': 0, 'functions': '利水渗湿，泄热', 'indications': '小便不利，水肿胀满'},
    '薏苡仁': {'category': '利水渗湿药', 'nature': '凉', 'taste': '甘、淡', 'meridian': '脾、胃、肺', 'min_dosage': 10, 'max_dosage': 30, 'is_toxic': 0, 'functions': '利水渗湿，健脾止泻，除痹排脓', 'indications': '水肿，脚气，脾虚泄泻'},
    
    # 温里药
    '附子': {'category': '温里药', 'nature': '大热', 'taste': '辛', 'meridian': '心、肾、脾', 'min_dosage': 3, 'max_dosage': 15, 'is_toxic': 1, 'functions': '回阳救逆，补火助阳，散寒止痛', 'indications': '亡阳虚脱，肢冷脉微，阳痿宫冷'},
    '干姜': {'category': '温里药', 'nature': '热', 'taste': '辛', 'meridian': '脾、胃、肾、心、肺', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '温中散寒，回阳通脉，温肺化饮', 'indications': '脘腹冷痛，呕吐泄泻，亡阳证'},
    '肉桂': {'category': '温里药', 'nature': '大热', 'taste': '辛、甘', 'meridian': '肾、脾、心、肝', 'min_dosage': 1, 'max_dosage': 5, 'is_toxic': 0, 'functions': '补火助阳，散寒止痛，温通经脉', 'indications': '阳痿宫冷，腰膝冷痛，心腹冷痛'},
    '吴茱萸': {'category': '温里药', 'nature': '热', 'taste': '辛、苦', 'meridian': '肝、脾、胃、肾', 'min_dosage': 1, 'max_dosage': 5, 'is_toxic': 1, 'functions': '散寒止痛，降逆止呕，助阳止泻', 'indications': '寒凝疼痛，胃寒呕吐'},
    
    # 理气药
    '陈皮': {'category': '理气药', 'nature': '温', 'taste': '辛、苦', 'meridian': '脾、肺', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '理气健脾，燥湿化痰', 'indications': '脘腹胀满，食少吐泻，咳嗽痰多'},
    '枳实': {'category': '理气药', 'nature': '微寒', 'taste': '苦、辛、酸', 'meridian': '脾、胃、大肠', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '破气消积，化痰散痞', 'indications': '积滞内停，痞满胀痛'},
    '木香': {'category': '理气药', 'nature': '温', 'taste': '辛、苦', 'meridian': '脾、胃、大肠、胆、三焦', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '行气止痛，健脾消食', 'indications': '脘腹胀痛，食积不消'},
    '香附': {'category': '理气药', 'nature': '平', 'taste': '辛、微苦、微甘', 'meridian': '肝、脾、三焦', 'min_dosage': 6, 'max_dosage': 10, 'is_toxic': 0, 'functions': '疏肝解郁，理气宽中，调经止痛', 'indications': '肝郁气滞，胸胁胀痛'},
    
    # 消食药
    '山楂': {'category': '消食药', 'nature': '微温', 'taste': '酸、甘', 'meridian': '脾、胃、肝', 'min_dosage': 10, 'max_dosage': 15, 'is_toxic': 0, 'functions': '消食健胃，行气散瘀', 'indications': '饮食积滞，泻痢腹痛'},
    '神曲': {'category': '消食药', 'nature': '温', 'taste': '甘、辛', 'meridian': '脾、胃', 'min_dosage': 6, 'max_dosage': 15, 'is_toxic': 0, 'functions': '消食和胃', 'indications': '饮食积滞'},
    
    # 止血药
    '地榆': {'category': '止血药', 'nature': '微寒', 'taste': '苦、酸、涩', 'meridian': '肝、大肠', 'min_dosage': 10, 'max_dosage': 15, 'is_toxic': 0, 'functions': '凉血止血，解毒敛疮', 'indications': '便血，痔血，水火烫伤'},
    
    # 活血化瘀药
    '川芎': {'category': '活血化瘀药', 'nature': '温', 'taste': '辛', 'meridian': '肝、胆、心包', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '活血行气，祛风止痛', 'indications': '血瘀气滞诸痛，头痛，风湿痹痛'},
    '丹参': {'category': '活血化瘀药', 'nature': '微寒', 'taste': '苦', 'meridian': '心、肝', 'min_dosage': 10, 'max_dosage': 15, 'is_toxic': 0, 'functions': '活血调经，祛瘀止痛，凉血消痈', 'indications': '月经不调，经闭痛经，疮疡肿痛'},
    '红花': {'category': '活血化瘀药', 'nature': '温', 'taste': '辛', 'meridian': '心、肝', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '活血通经，散瘀止痛', 'indications': '经闭，痛经，跌打损伤'},
    '桃仁': {'category': '活血化瘀药', 'nature': '平', 'taste': '苦、甘', 'meridian': '心、肝、大肠', 'min_dosage': 5, 'max_dosage': 10, 'is_toxic': 1, 'functions': '活血祛瘀，润肠通便', 'indications': '经闭痛经，癥瘕痞块，肠燥便秘'},
    '延胡索': {'category': '活血化瘀药', 'nature': '温', 'taste': '辛、苦', 'meridian': '心、肝、脾', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '活血，行气，止痛', 'indications': '气血瘀滞诸痛'},
    '莪术': {'category': '活血化瘀药', 'nature': '温', 'taste': '辛、苦', 'meridian': '肝、脾', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '破血行气，消积止痛', 'indications': '癥瘕痞块，瘀血经闭'},
    
    # 化痰止咳平喘药
    '半夏': {'category': '化痰止咳平喘药', 'nature': '温', 'taste': '辛', 'meridian': '脾、胃、肺', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 1, 'functions': '燥湿化痰，降逆止呕，消痞散结', 'indications': '湿痰寒痰，呕吐，心下痞'},
    '天南星': {'category': '化痰止咳平喘药', 'nature': '温', 'taste': '苦、辛', 'meridian': '肺、肝、脾', 'min_dosage': 3, 'max_dosage': 9, 'is_toxic': 1, 'functions': '燥湿化痰，祛风解痉', 'indications': '顽痰咳嗽，风痰眩晕'},
    '桔梗': {'category': '化痰止咳平喘药', 'nature': '平', 'taste': '苦、辛', 'meridian': '肺', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '宣肺祛痰，利咽排脓', 'indications': '咳嗽痰多，咽喉肿痛，肺痈'},
    '杏仁': {'category': '化痰止咳平喘药', 'nature': '微温', 'taste': '苦', 'meridian': '肺、大肠', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 1, 'functions': '降气止咳平喘，润肠通便', 'indications': '咳嗽气喘，肠燥便秘'},
    '款冬花': {'category': '化痰止咳平喘药', 'nature': '温', 'taste': '辛', 'meridian': '肺', 'min_dosage': 5, 'max_dosage': 10, 'is_toxic': 0, 'functions': '润肺下气，止咳化痰', 'indications': '咳嗽气喘'},
    '紫菀': {'category': '化痰止咳平喘药', 'nature': '温', 'taste': '辛、苦', 'meridian': '肺', 'min_dosage': 5, 'max_dosage': 10, 'is_toxic': 0, 'functions': '润肺下气，消痰止咳', 'indications': '咳嗽痰多'},
    '贝母': {'category': '化痰止咳平喘药', 'nature': '微寒', 'taste': '苦、甘', 'meridian': '肺、心', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '清热化痰，润肺止咳，散结消肿', 'indications': '虚劳咳嗽，肺热燥咳'},
    '川贝母': {'category': '化痰止咳平喘药', 'nature': '微寒', 'taste': '苦、甘', 'meridian': '肺、心', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '清热化痰，润肺止咳', 'indications': '肺热燥咳，干咳少痰'},
    '瓜蒌': {'category': '化痰止咳平喘药', 'nature': '寒', 'taste': '甘、微苦', 'meridian': '肺、胃、大肠', 'min_dosage': 10, 'max_dosage': 20, 'is_toxic': 0, 'functions': '清热化痰，宽胸散结，润肠通便', 'indications': '痰热咳嗽，胸痹心痛'},
    
    # 安神药
    '酸枣仁': {'category': '安神药', 'nature': '平', 'taste': '甘、酸', 'meridian': '肝、胆、心', 'min_dosage': 10, 'max_dosage': 20, 'is_toxic': 0, 'functions': '养心补肝，宁心安神，敛汗生津', 'indications': '虚烦不眠，惊悸多梦'},
    '远志': {'category': '安神药', 'nature': '温', 'taste': '苦、辛', 'meridian': '心、肾、肺', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '安神益智，祛痰开窍', 'indications': '失眠多梦，健忘惊悸'},
    '龙骨': {'category': '安神药', 'nature': '平', 'taste': '甘、涩', 'meridian': '心、肝、肾', 'min_dosage': 15, 'max_dosage': 30, 'is_toxic': 0, 'functions': '镇惊安神，平肝潜阳', 'indications': '心神不宁，心悸失眠'},
    '牡蛎': {'category': '安神药', 'nature': '微寒', 'taste': '咸', 'meridian': '肝、胆、肾', 'min_dosage': 10, 'max_dosage': 30, 'is_toxic': 0, 'functions': '重镇安神，潜阳补阴，软坚散结', 'indications': '惊悸失眠，眩晕耳鸣'},
    
    # 平肝息风药
    '天麻': {'category': '平肝息风药', 'nature': '平', 'taste': '甘', 'meridian': '肝', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '息风止痉，平抑肝阳，祛风通络', 'indications': '眩晕头痛，惊痫抽搐'},
    '钩藤': {'category': '平肝息风药', 'nature': '凉', 'taste': '甘', 'meridian': '肝、心包', 'min_dosage': 3, 'max_dosage': 12, 'is_toxic': 0, 'functions': '清热平肝，息风止痉', 'indications': '头痛眩晕，惊痫抽搐'},
    '全蝎': {'category': '平肝息风药', 'nature': '平', 'taste': '辛', 'meridian': '肝', 'min_dosage': 2, 'max_dosage': 5, 'is_toxic': 1, 'functions': '息风镇痉，通络止痛，攻毒散结', 'indications': '痉挛抽搐，疮疡肿毒'},
    
    # 开窍药
    '麝香': {'category': '开窍药', 'nature': '温', 'taste': '辛', 'meridian': '心、脾', 'min_dosage': 0.03, 'max_dosage': 0.1, 'is_toxic': 0, 'functions': '开窍醒神，活血通经，消肿止痛', 'indications': '热病神昏，中风痰厥'},
    
    # 补虚药
    '人参': {'category': '补虚药', 'nature': '微温', 'taste': '甘、微苦', 'meridian': '脾、肺、心、肾', 'min_dosage': 3, 'max_dosage': 9, 'is_toxic': 0, 'functions': '大补元气，复脉固脱，补脾益肺', 'indications': '体虚欲脱，肢冷脉微，脾虚食少'},
    '党参': {'category': '补虚药', 'nature': '平', 'taste': '甘', 'meridian': '脾、肺', 'min_dosage': 9, 'max_dosage': 30, 'is_toxic': 0, 'functions': '补中益气，健脾益肺', 'indications': '脾肺虚弱，气短心悸，食少便溏'},
    '黄芪': {'category': '补虚药', 'nature': '微温', 'taste': '甘', 'meridian': '脾、肺', 'min_dosage': 10, 'max_dosage': 30, 'is_toxic': 0, 'functions': '补气升阳，固表止汗，利水消肿', 'indications': '气虚乏力，食少便溏，中气下陷'},
    '白术': {'category': '补虚药', 'nature': '温', 'taste': '甘、苦', 'meridian': '脾、胃', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '健脾益气，燥湿利水，止汗', 'indications': '脾虚食少，腹胀泄泻，痰饮眩悸'},
    '山药': {'category': '补虚药', 'nature': '平', 'taste': '甘', 'meridian': '脾、肺、肾', 'min_dosage': 15, 'max_dosage': 30, 'is_toxic': 0, 'functions': '补脾养胃，生津益肺，补肾涩精', 'indications': '脾虚食少，久泻不止，肺虚喘咳'},
    '甘草': {'category': '补虚药', 'nature': '平', 'taste': '甘', 'meridian': '心、肺、脾、胃', 'min_dosage': 2, 'max_dosage': 10, 'is_toxic': 0, 'functions': '补脾益气，清热解毒，祛痰止咳', 'indications': '脾胃虚弱，倦怠乏力，心悸气短'},
    '当归': {'category': '补虚药', 'nature': '温', 'taste': '甘、辛', 'meridian': '肝、心、脾', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '补血活血，调经止痛，润肠通便', 'indications': '血虚萎黄，眩晕心悸，月经不调'},
    '熟地黄': {'category': '补虚药', 'nature': '微温', 'taste': '甘', 'meridian': '肝、肾', 'min_dosage': 10, 'max_dosage': 30, 'is_toxic': 0, 'functions': '补血滋阴，益精填髓', 'indications': '血虚萎黄，心悸怔忡，月经不调'},
    '白芍': {'category': '补虚药', 'nature': '微寒', 'taste': '苦、酸', 'meridian': '肝、脾', 'min_dosage': 6, 'max_dosage': 15, 'is_toxic': 0, 'functions': '养血调经，敛阴止汗，柔肝止痛', 'indications': '血虚萎黄，月经不调，自汗盗汗'},
    '阿胶': {'category': '补虚药', 'nature': '平', 'taste': '甘', 'meridian': '肺、肝、肾', 'min_dosage': 3, 'max_dosage': 9, 'is_toxic': 0, 'functions': '补血滋阴，润燥止血', 'indications': '血虚萎黄，眩晕心悸，多种出血证'},
    '麦冬': {'category': '补虚药', 'nature': '微寒', 'taste': '甘、微苦', 'meridian': '心、肺、胃', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '养阴润肺，益胃生津，清心除烦', 'indications': '肺燥干咳，阴虚痨嗽，喉痹咽痛'},
    '枸杞子': {'category': '补虚药', 'nature': '平', 'taste': '甘', 'meridian': '肝、肾', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '滋补肝肾，益精明目', 'indications': '虚劳精亏，腰膝酸痛，眩晕耳鸣'},
    '菟丝子': {'category': '补虚药', 'nature': '平', 'taste': '辛、甘', 'meridian': '肝、肾、脾', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '补肾益精，养肝明目', 'indications': '阳痿遗精，尿有余沥，目昏耳鸣'},
    '杜仲': {'category': '补虚药', 'nature': '温', 'taste': '甘', 'meridian': '肝、肾', 'min_dosage': 10, 'max_dosage': 15, 'is_toxic': 0, 'functions': '补肝肾，强筋骨，安胎', 'indications': '肾虚腰痛，筋骨无力，妊娠漏血'},
    '续断': {'category': '补虚药', 'nature': '微温', 'taste': '苦、辛', 'meridian': '肝、肾', 'min_dosage': 9, 'max_dosage': 15, 'is_toxic': 0, 'functions': '补肝肾，强筋骨，续折伤', 'indications': '腰膝酸软，风湿痹痛，跌打损伤'},
    '肉苁蓉': {'category': '补虚药', 'nature': '温', 'taste': '甘、咸', 'meridian': '肾、大肠', 'min_dosage': 10, 'max_dosage': 15, 'is_toxic': 0, 'functions': '补肾阳，益精血，润肠通便', 'indications': '阳痿不孕，腰膝酸软，肠燥便秘'},
    '淫羊藿': {'category': '补虚药', 'nature': '温', 'taste': '辛、甘', 'meridian': '肝、肾', 'min_dosage': 6, 'max_dosage': 10, 'is_toxic': 0, 'functions': '补肾阳，强筋骨，祛风湿', 'indications': '阳痿遗精，筋骨痿软，风湿痹痛'},
    '补骨脂': {'category': '补虚药', 'nature': '温', 'taste': '辛、苦', 'meridian': '肾、脾', 'min_dosage': 6, 'max_dosage': 10, 'is_toxic': 0, 'functions': '补肾壮阳，固精缩尿，温脾止泻', 'indications': '腰膝冷痛，阳痿遗精，遗尿尿频'},
    '沙参': {'category': '补虚药', 'nature': '微寒', 'taste': '甘、微苦', 'meridian': '肺、胃', 'min_dosage': 10, 'max_dosage': 15, 'is_toxic': 0, 'functions': '养阴清肺，益胃生津', 'indications': '肺热燥咳，阴虚劳嗽，胃阴不足'},
    '石斛': {'category': '补虚药', 'nature': '微寒', 'taste': '甘', 'meridian': '胃、肾', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '益胃生津，滋阴清热', 'indications': '热病津伤，口干烦渴，胃阴不足'},
    '百合': {'category': '补虚药', 'nature': '微寒', 'taste': '甘', 'meridian': '心、肺', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '养阴润肺，清心安神', 'indications': '阴虚燥咳，劳嗽咳血，虚烦惊悸'},
    '黄精': {'category': '补虚药', 'nature': '平', 'taste': '甘', 'meridian': '脾、肺、肾', 'min_dosage': 9, 'max_dosage': 15, 'is_toxic': 0, 'functions': '补气养阴，健脾润肺，益肾', 'indications': '脾胃气虚，肺虚燥咳，精血不足'},
    '大枣': {'category': '补虚药', 'nature': '温', 'taste': '甘', 'meridian': '脾、胃、心', 'min_dosage': 6, 'max_dosage': 15, 'is_toxic': 0, 'functions': '补中益气，养血安神', 'indications': '脾虚食少，乏力便溏，妇人脏躁'},
    '何首乌': {'category': '补虚药', 'nature': '微温', 'taste': '苦、甘、涩', 'meridian': '肝、肾', 'min_dosage': 3, 'max_dosage': 6, 'is_toxic': 1, 'functions': '制用补肝肾益精血，生用解毒截疟', 'indications': '血虚萎黄，须发早白，久疟'},
    
    # 收涩药
    '五味子': {'category': '收涩药', 'nature': '温', 'taste': '酸、甘', 'meridian': '肺、心、肾', 'min_dosage': 2, 'max_dosage': 6, 'is_toxic': 0, 'functions': '收敛固涩，益气生津，补肾宁心', 'indications': '久嗽虚喘，梦遗滑精，遗尿尿频'},
    '山茱萸': {'category': '收涩药', 'nature': '微温', 'taste': '酸、涩', 'meridian': '肝、肾', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '补益肝肾，收涩固脱', 'indications': '眩晕耳鸣，腰膝酸痛，阳痿遗精'},
    '乌梅': {'category': '收涩药', 'nature': '平', 'taste': '酸、涩', 'meridian': '肝、脾、肺、大肠', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '敛肺止咳，涩肠止泻，安蛔止痛', 'indications': '肺虚久咳，久泻久痢，蛔厥腹痛'},
    '莲子': {'category': '收涩药', 'nature': '平', 'taste': '甘、涩', 'meridian': '脾、肾、心', 'min_dosage': 6, 'max_dosage': 15, 'is_toxic': 0, 'functions': '补脾止泻，止带，益肾涩精', 'indications': '脾虚泄泻，带下，遗精滑精'},
    
    # 其他常见药（从 Excel 中提取的部分药材）
    '牛膝': {'category': '活血化瘀药', 'nature': '平', 'taste': '苦、甘、酸', 'meridian': '肝、肾', 'min_dosage': 6, 'max_dosage': 15, 'is_toxic': 0, 'functions': '逐瘀通经，补肝肾强筋骨', 'indications': '经闭痛经，腰膝酸痛，筋骨无力'},
    '柴胡': {'category': '解表药', 'nature': '微寒', 'taste': '辛、苦', 'meridian': '肝、胆', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '疏散退热，疏肝解郁，升举阳气', 'indications': '感冒发热，寒热往来，胸胁胀痛'},
    '升麻': {'category': '解表药', 'nature': '微寒', 'taste': '辛、微甘', 'meridian': '肺、脾、胃、大肠', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '发表透疹，清热解毒，升举阳气', 'indications': '风热感冒，麻疹不透，齿痛口疮'},
    '葛根': {'category': '解表药', 'nature': '凉', 'taste': '甘、辛', 'meridian': '脾、胃、肺', 'min_dosage': 10, 'max_dosage': 15, 'is_toxic': 0, 'functions': '解肌退热，生津止渴，透疹', 'indications': '外感发热，项背强痛，口渴'},
    '淡豆豉': {'category': '解表药', 'nature': '凉', 'taste': '辛、甘、微苦', 'meridian': '肺、胃', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '解表，除烦', 'indications': '感冒，寒热头痛，心烦'},
    '栀子': {'category': '清热药', 'nature': '寒', 'taste': '苦', 'meridian': '心、肺、三焦', 'min_dosage': 6, 'max_dosage': 10, 'is_toxic': 0, 'functions': '泻火除烦，清热利湿，凉血解毒', 'indications': '热病心烦，湿热黄疸，血热吐衄'},
    '夏枯草': {'category': '清热药', 'nature': '寒', 'taste': '辛、苦', 'meridian': '肝、胆', 'min_dosage': 9, 'max_dosage': 15, 'is_toxic': 0, 'functions': '清热泻火，明目，散结消肿', 'indications': '目赤肿痛，头痛眩晕，瘰疬瘿瘤'},
    '竹叶': {'category': '清热药', 'nature': '寒', 'taste': '甘、辛、淡', 'meridian': '心、胃、小肠', 'min_dosage': 6, 'max_dosage': 15, 'is_toxic': 0, 'functions': '清热泻火，除烦利尿', 'indications': '热病烦渴，小便短赤'},
    '淡竹叶': {'category': '清热药', 'nature': '寒', 'taste': '甘、淡', 'meridian': '心、胃、小肠', 'min_dosage': 6, 'max_dosage': 10, 'is_toxic': 0, 'functions': '清热泻火，除烦利尿', 'indications': '热病烦渴，口疮尿赤'},
    '龙胆': {'category': '清热药', 'nature': '寒', 'taste': '苦', 'meridian': '肝、胆', 'min_dosage': 3, 'max_dosage': 6, 'is_toxic': 0, 'functions': '清热燥湿，泻肝胆火', 'indications': '湿热黄疸，阴肿阴痒，目赤'},
    '蒲公英': {'category': '清热药', 'nature': '寒', 'taste': '苦、甘', 'meridian': '肝、胃', 'min_dosage': 10, 'max_dosage': 15, 'is_toxic': 0, 'functions': '清热解毒，消肿散结，利湿通淋', 'indications': '痈肿疔毒，乳痈内痈，热淋涩痛'},
    '板蓝根': {'category': '清热药', 'nature': '寒', 'taste': '苦', 'meridian': '心、胃', 'min_dosage': 10, 'max_dosage': 15, 'is_toxic': 0, 'functions': '清热解毒，凉血利咽', 'indications': '温疫时毒，发热咽痛'},
    '白茅根': {'category': '止血药', 'nature': '寒', 'taste': '甘', 'meridian': '肺、胃、膀胱', 'min_dosage': 10, 'max_dosage': 30, 'is_toxic': 0, 'functions': '凉血止血，清热利尿', 'indications': '血热吐血，衄血，尿血'},
    '三七': {'category': '止血药', 'nature': '温', 'taste': '甘、微苦', 'meridian': '肝、胃', 'min_dosage': 3, 'max_dosage': 9, 'is_toxic': 0, 'functions': '散瘀止血，消肿定痛', 'indications': '咯血，吐血，衄血，便血，跌打损伤'},
    '蒲黄': {'category': '止血药', 'nature': '平', 'taste': '甘', 'meridian': '肝、心包', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '止血，化瘀，通淋', 'indications': '吐血，衄血，咯血，崩漏'},
    '茜草': {'category': '止血药', 'nature': '寒', 'taste': '苦', 'meridian': '肝', 'min_dosage': 6, 'max_dosage': 10, 'is_toxic': 0, 'functions': '凉血，祛瘀，止血，通经', 'indications': '吐血，衄血，崩漏，经闭瘀阻'},
    '益母草': {'category': '活血化瘀药', 'nature': '微寒', 'taste': '苦、辛', 'meridian': '肝、心包、膀胱', 'min_dosage': 10, 'max_dosage': 30, 'is_toxic': 0, 'functions': '活血调经，利尿消肿', 'indications': '月经不调，痛经经闭，恶露不尽'},
    '鸡血藤': {'category': '活血化瘀药', 'nature': '温', 'taste': '苦、甘', 'meridian': '肝、肾', 'min_dosage': 10, 'max_dosage': 30, 'is_toxic': 0, 'functions': '活血补血，调经止痛，舒筋活络', 'indications': '月经不调，痛经经闭，风湿痹痛'},
    '郁金': {'category': '活血化瘀药', 'nature': '寒', 'taste': '辛、苦', 'meridian': '肝、心、肺', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '活血止痛，行气解郁，清心凉血', 'indications': '胸胁刺痛，胸痹心痛，热病神昏'},
    '姜黄': {'category': '活血化瘀药', 'nature': '温', 'taste': '辛、苦', 'meridian': '肝、脾', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '破血行气，通经止痛', 'indications': '胸胁刺痛，胸痹心痛'},
    '穿山甲': {'category': '活血化瘀药', 'nature': '微寒', 'taste': '咸', 'meridian': '肝、胃', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '活血消癥，通经下乳，消肿排脓', 'indications': '癥瘕痞块，瘀血经闭'},
    '没药': {'category': '活血化瘀药', 'nature': '平', 'taste': '辛、苦', 'meridian': '心、肝、脾', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '散瘀定痛，消肿生肌', 'indications': '胸痹心痛，胃脘疼痛，跌打损伤'},
    '乳香': {'category': '活血化瘀药', 'nature': '温', 'taste': '辛、苦', 'meridian': '心、肝、脾', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '活血定痛，消肿生肌', 'indications': '胸痹心痛，胃脘疼痛，跌打损伤'},
    '三棱': {'category': '活血化瘀药', 'nature': '平', 'taste': '辛、苦', 'meridian': '肝、脾', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '破血行气，消积止痛', 'indications': '癥瘕痞块，瘀血经闭'},
    
    # 更多常用药
    '前胡': {'category': '化痰止咳平喘药', 'nature': '微寒', 'taste': '苦、辛', 'meridian': '肺', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '降气化痰，散风清热', 'indications': '痰热喘满，咯痰黄稠'},
    '桑白皮': {'category': '化痰止咳平喘药', 'nature': '寒', 'taste': '甘', 'meridian': '肺', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '泻肺平喘，利水消肿', 'indications': '肺热喘咳，水肿胀满'},
    '葶苈子': {'category': '化痰止咳平喘药', 'nature': '大寒', 'taste': '辛、苦', 'meridian': '肺、膀胱', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '泻肺平喘，行水消肿', 'indications': '痰涎壅肺，喘咳痰多'},
    '枇杷叶': {'category': '化痰止咳平喘药', 'nature': '微寒', 'taste': '苦', 'meridian': '肺、胃', 'min_dosage': 6, 'max_dosage': 10, 'is_toxic': 0, 'functions': '清肺止咳，降逆止呕', 'indications': '肺热咳嗽，气逆喘急'},
    '旋覆花': {'category': '化痰止咳平喘药', 'nature': '微温', 'taste': '苦、辛、咸', 'meridian': '肺、脾、胃、大肠', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '降气，消痰，行水，止呕', 'indications': '风寒咳嗽，痰饮蓄结'},
    '白前': {'category': '化痰止咳平喘药', 'nature': '微温', 'taste': '辛、苦', 'meridian': '肺', 'min_dosage': 3, 'max_dosage': 10, 'is_toxic': 0, 'functions': '降气，消痰，止咳', 'indications': '肺气壅实，咳嗽痰多'},
    '竹茹': {'category': '化痰止咳平喘药', 'nature': '微寒', 'taste': '甘', 'meridian': '肺、胃、心、胆', 'min_dosage': 6, 'max_dosage': 10, 'is_toxic': 0, 'functions': '清热化痰，除烦止呕', 'indications': '痰热咳嗽，胆火挟痰'},
    '海藻': {'category': '化痰止咳平喘药', 'nature': '寒', 'taste': '苦、咸', 'meridian': '肝、胃、肾', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '消痰软坚散结，利水消肿', 'indications': '瘿瘤，瘰疬，睾丸肿痛'},
    '昆布': {'category': '化痰止咳平喘药', 'nature': '寒', 'taste': '咸', 'meridian': '肝、胃、肾', 'min_dosage': 6, 'max_dosage': 12, 'is_toxic': 0, 'functions': '消痰软坚散结，利水消肿', 'indications': '瘿瘤，瘰疬，睾丸肿痛'},
}

print(f"\n内置了 {len(TCM_HERBS)} 种药材的完整属性")

# ============ 3. 获取 Excel 中所有药材名 ============
print(f"\nExcel 中的药材: {sorted(herbs_in_excel)}")

# 看看哪些药材没有内置属性
missing_herbs = [h for h in herbs_in_excel if h not in TCM_HERBS]
if missing_herbs:
    print(f"\n⚠️ 以下 {len(missing_herbs)} 种药材没有内置属性，将使用基本属性补充:")
    for h in missing_herbs:
        print(f"  - {h}")
else:
    print("\n✅ Excel 中所有药材都已包含内置属性")

# 补充缺失的药材（使用基本属性）
for h in missing_herbs:
    # 清洗药材名
    clean_name = h.strip()
    if len(clean_name) >= 2:
        TCM_HERBS[clean_name] = {
            'category': '其他', 
            'nature': '平', 
            'taste': '', 
            'meridian': '', 
            'min_dosage': 3, 
            'max_dosage': 10, 
            'is_toxic': 0, 
            'functions': '', 
            'indications': ''
        }

print(f"\n补充后共有 {len(TCM_HERBS)} 种药材")

# ============ 4. 写入 herb_data.sql ============
print("\n正在生成 herb_data.sql ...")
with open(f"{OUTPUT_DIR}/herb_data.sql", 'w', encoding='utf-8') as f:
    f.write("-- ============================================================\n")
    f.write("-- 中药材基础数据导入\n")
    f.write("-- 来源: 中医医疗系统 + 标注二级分类后的方剂表.xlsx\n")
    f.write("-- ============================================================\n\n")
    f.write("USE tcm_hospital;\n\n")
    f.write("-- 清空已有数据（避免重复导入）\n")
    f.write("SET FOREIGN_KEY_CHECKS = 0;\n")
    f.write("TRUNCATE TABLE herb;\n")
    f.write("SET FOREIGN_KEY_CHECKS = 1;\n\n")
    
    for name, info in sorted(TCM_HERBS.items()):
        functions = info['functions'].replace("'", "''")
        indications = info['indications'].replace("'", "''")
        f.write(f"INSERT INTO herb (name, category, nature, taste, meridian, min_dosage, max_dosage, unit, is_toxic, functions, indications) VALUES\n")
        f.write(f"('{name}', '{info['category']}', '{info['nature']}', '{info['taste']}', '{info['meridian']}', {info['min_dosage']}, {info['max_dosage']}, 'g', {info['is_toxic']}, '{functions}', '{indications}');\n")

print("✅ herb_data.sql 生成完成")

# ============ 5. 写入 formula_data.sql ============
print("正在生成 formula_data.sql ...")

# 为每个方剂生成 herb_name -> id 的映射
# 先用方剂名去重
formula_list = list(formulas.values())

with open(f"{OUTPUT_DIR}/formula_data.sql", 'w', encoding='utf-8') as f:
    f.write("-- ============================================================\n")
    f.write("-- 方剂数据导入\n")
    f.write("-- 来源: 标注二级分类后的方剂表.xlsx\n")
    f.write("-- ============================================================\n\n")
    f.write("USE tcm_hospital;\n\n")
    f.write("-- 清空已有数据\n")
    f.write("SET FOREIGN_KEY_CHECKS = 0;\n")
    f.write("TRUNCATE TABLE formula_herb;\n")
    f.write("TRUNCATE TABLE formula;\n")
    f.write("SET FOREIGN_KEY_CHECKS = 1;\n\n")
    
    # INSERT formulas
    f.write("-- 插入方剂数据\n")
    for fm in formula_list:
        name = fm['name'].replace("'", "''")
        func = fm['functions'].replace("'", "''")
        ind = fm['indications'].replace("'", "''")
        mod = fm['modifications'].replace("'", "''")
        src = fm['source'].replace("'", "''")
        cat_l1 = fm['category_l1'].replace("'", "''")
        cat_l2 = fm['category_l2'].replace("'", "''")
        
        f.write(f"INSERT INTO formula (name, category_l1, category_l2, functions, indications, modifications, source) VALUES\n")
        f.write(f"('{name}', '{cat_l1}', '{cat_l2}', '{func}', '{ind}', '{mod}', '{src}');\n")
    
    f.write("\n")
    
    # INSERT formula_herb
    f.write("-- 插入方剂-药材关联数据\n")
    f.write("-- 假设: herb.name 唯一, formula.name 唯一\n")
    for fh in formula_herbs:
        fm_name = fh['formula_name'].replace("'", "''")
        hb_name = fh['herb_name'].replace("'", "''")
        dosage = fh['dosage'].replace("'", "''")
        
        f.write(f"INSERT INTO formula_herb (formula_id, herb_id, dosage)\n")
        f.write(f"SELECT f.id, h.id, '{dosage}'\n")
        f.write(f"FROM formula f, herb h\n")
        f.write(f"WHERE f.name = '{fm_name}' AND h.name = '{hb_name}';\n")

print("✅ formula_data.sql 生成完成")

# ============ 6. 写入 symptom_dict_data.sql ============
print("正在生成 symptom_dict_data.sql ...")
symptoms_data = {
    'chief_complaint': [
        ('头痛', 1), ('眩晕', 1), ('咳嗽', 1), ('发热', 1), ('腹痛', 1),
        ('胸闷', 1), ('心悸', 1), ('失眠', 1), ('呕吐', 1), ('腹泻', 1),
        ('便秘', 1), ('腰酸', 1), ('乏力', 1), ('气喘', 1), ('咽痛', 1),
        ('胃痛', 1), ('水肿', 1), ('耳鸣', 1), ('出汗', 0), ('口干', 1),
        ('鼻塞', 0), ('关节痛', 0), ('尿频', 0), ('月经不调', 0), ('胁痛', 0)
    ],
    'present_illness': {
        '诱因': [('受凉', 1), ('劳累', 1), ('饮食不节', 1), ('情志刺激', 1), ('外伤', 0)],
        '特征': [('持续性', 1), ('阵发性', 1), ('进行性加重', 0), ('反复发作', 1), ('夜间加重', 0)],
        '伴随': [('恶寒发热', 1), ('自汗盗汗', 0), ('恶心呕吐', 0), ('食欲不振', 1), ('口渴引饮', 0)]
    },
    'past_history': [
        ('高血压', 1), ('糖尿病', 1), ('心脏病', 1), ('肝炎', 0), ('结核', 0),
        ('手术史', 0), ('外伤史', 0), ('过敏史', 1)
    ],
    'allergy': [
        ('青霉素过敏', 1), ('磺胺类过敏', 1), ('花粉过敏', 0), ('食物过敏', 0)
    ],
    'personal': [
        ('吸烟', 1), ('饮酒', 1), ('喜食辛辣', 0), ('喜食生冷', 0),
        ('久坐', 0), ('熬夜', 1), ('压力大', 1)
    ]
}

with open(f"{OUTPUT_DIR}/symptom_dict_data.sql", 'w', encoding='utf-8') as f:
    f.write("-- ============================================================\n")
    f.write("-- 症状/标签字典数据导入\n")
    f.write("-- ============================================================\n\n")
    f.write("USE tcm_hospital;\n\n")
    f.write("SET FOREIGN_KEY_CHECKS = 0;\n")
    f.write("TRUNCATE TABLE symptom_dict;\n")
    f.write("SET FOREIGN_KEY_CHECKS = 1;\n\n")
    
    sort = 0
    for category, items in symptoms_data.items():
        if isinstance(items, list):
            for label, is_common in items:
                sort += 1
                w = 5 if is_common else 0
                f.write(f"INSERT INTO symptom_dict (category, label, sort_order, is_common, weight) VALUES\n")
                f.write(f"('{category}', '{label}', {sort}, {is_common}, {w});\n")
        elif isinstance(items, dict):
            for sub_category, sub_items in items.items():
                for label, is_common in sub_items:
                    sort += 1
                    w = 5 if is_common else 0
                    f.write(f"INSERT INTO symptom_dict (category, sub_category, label, sort_order, is_common, weight) VALUES\n")
                    f.write(f"('{category}', '{sub_category}', '{label}', {sort}, {is_common}, {w});\n")

print("✅ symptom_dict_data.sql 生成完成")

# ============ 7. 总统计 ============
print("\n" + "=" * 50)
print("📊 数据导入统计")
print("=" * 50)
print(f"  药材数据: {len(TCM_HERBS)} 条")
print(f"  方剂数据: {len(formulas)} 条")
print(f"  方剂-药材关联: {len(formula_herbs)} 条")
print(f"  症状字典: {sort} 条")
print("\n📁 生成的文件:")
print(f"  1. {OUTPUT_DIR}/herb_data.sql")
print(f"  2. {OUTPUT_DIR}/formula_data.sql")
print(f"  3. {OUTPUT_DIR}/symptom_dict_data.sql")
print("\n💡 在 Navicat 中按以下顺序执行:")
print("  1. symptom_dict_data.sql (症状字典)")
print("  2. herb_data.sql (药材数据)")
print("  3. formula_data.sql (方剂数据)\n")